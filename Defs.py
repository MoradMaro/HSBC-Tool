from calendar import c
from errno import ENOMSG
from zlib import DEF_BUF_SIZE
import streamlit as st
from sqlalchemy import create_engine, text
import pandas as pd
from io import BytesIO
import numpy as np
import re
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from dateutil.parser import parse
from datetime import date

# إعدادات الاتصال بقاعدة البيانات
DB_Host = "localhost"
DB_Name = "HSBC"
DB_User = "postgres"
DB_Password = "MoradTawfik123456789##"
DB_Port = "5432"

# -----------------------------------------------------------
# 0- إنشاء الاتصال بقاعدة البيانات (مرة واحدة فقط)
@st.cache_resource
def get_db_engine():
    return create_engine(f'postgresql+psycopg2://{DB_User}:{DB_Password}@{DB_Host}:{DB_Port}/{DB_Name}')

# -----------------------------------------------------------
# 1- Clear Table (غير مناسب للتخزين المؤقت لأنه يعدّل البيانات)
# @st.cache_data(ttl=600)
def clear_table(table_name):
    try:
        engine = get_db_engine()
        with engine.connect() as conn:
            count = conn.execute(text(f'SELECT COUNT(*) FROM "{table_name}"')).scalar()
            if not count:
                return f'The table {table_name} is already empty.'
            conn.execute(text(f'TRUNCATE TABLE "{table_name}" RESTART IDENTITY CASCADE'))
            conn.commit()
            return f'Data Cleared from {table_name}'
    except Exception as e:
        return f"Error: {str(e)}"

# -----------------------------------------------------------
# 2- Load File to Table (غير مناسب للتخزين المؤقت لأنه يُحمّل بيانات)

# @st.cache_data(ttl=600)
# def load_file_to_table(df, table_name):
#     try:
#         df.columns = df.columns.str.strip()
#         if df.empty or df.columns.size == 0:
#             return "⚠️ The uploaded data is empty or contains no valid columns."
#         engine = get_db_engine()
#         with engine.connect() as conn:
#             count = conn.execute(text(f'SELECT COUNT(*) FROM "{table_name}"')).scalar()
#             if count is not None and count > 0:
#                 return f"⚠️ The table '{table_name}' already contains data ({count} rows). Data loading aborted."
#         df.to_sql(table_name, con=engine, if_exists='append', index=False)
#         return f"✅ Data successfully loaded into '{table_name}' ({len(df)} rows)."
#     except Exception as e:
#         return f"❌ Error: {str(e)}"

# @st.cache_data(ttl=600)
def load_file_to_table(df, table_name):
    try:
        # ✅ تعيين أسماء الأعمدة يدويًا لجدول DFF فقط
        if table_name == "DFF":
            custom_columns = [
                'cashp_id', 'cp_type', 'cycle_type', 'crncy_id', 'crcy_typ', 'comp_id',
                'denom_id', 'date', 'cassette', 'open_bal', 'nopen_bal', 'norm_del', 'nnorm_del',
                'norm_rtr', 'nnorm_rtr', 'unpl_del', 'nunpl_del', 'unpl_rtr', 'nunpl_rtr',
                'with_tran', 'wthdrwls', 'nwthdrwls', 'pre_srv', 'npre_srv', 'dep_tran',
                'deposits', 'ndeposits', 'clos_bal', 'nclos_bal', 'bal_disp', 'bal_escr',
                'bal_unav', 'opr_stat', 'excld_fl', '1', 'date2'
            ]
            if df.shape[1] != len(custom_columns):
                return f"❌ Column count mismatch: expected {len(custom_columns)} columns, got {df.shape[1]}"
            df.columns = custom_columns
        else:
            df.columns = df.columns.str.strip()

        if df.empty or len(df.columns) == 0:
            return "⚠️ The uploaded data is empty or contains no valid columns."

        engine = get_db_engine()
        with engine.connect() as conn:
            count = conn.execute(text(f'SELECT COUNT(*) FROM "{table_name}"')).scalar()
            if count is not None and count > 0:
                return f"⚠️ The table '{table_name}' already contains data ({count} rows). Data loading aborted."

        df.to_sql(table_name, con=engine, if_exists='append', index=False)
        return f"✅ Data successfully loaded into '{table_name}' ({len(df)} rows)."

    except Exception as e:
        return f"❌ Error: {str(e)}"



#&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&

#  حساب أيام تغطية الكاسيت (Cassette Cover Days) من DataFrame
def cassette_cover_days(df): 
    with np.errstate(divide='ignore', invalid='ignore'):
        cover_days = (df["nclos_bal"] / df["nwthdrwls"]).replace([np.inf, -np.inf], np.nan).round(0)
    return cover_days.astype('Int64')  # عمود رقمي مع NaN
# -----------------------------------------------------------

# تحويل القيم NaN إلى سلسلة نصية "ـــــ" في عمود DataFrame 
def cover_nan_to_str(series): 
    # return series.apply(lambda x: int(x) if pd.notna(x) else "ـــــ")
    return series.apply(
        lambda x: f"({abs(int(x))})" if pd.notna(x) and x < 0 else (str(int(x)) if pd.notna(x) else "ـــــ")
    )

# -----------------------------------------------------------
# حساب الرصيد اليومي


def Balance_Today(df):
    df["Balance Today"] = (df["nclos_bal"] - df["nwthdrwls"]).astype("Int64")
    return df.sort_values(by="Balance Today", ascending=True)
# -----------------------------------------------------------
# حساب الرصيد غدًا
def Balance_Tomorrow(df):
    df["Balance Tomorrow"] = (df["Balance Today"] - df["nwthdrwls"]).astype("Int64")
    return df.sort_values(by="Balance Tomorrow", ascending=True)

# -----------------------------------------------------------

# Get Needs order today
def get_needs_order_today(df):
    # تأكد أن العمودين موجودان
    if "denom_id" not in df.columns or "Balance Today" not in df.columns:
        st.error("❌ الأعمدة 'denom_id' أو 'Balance Today' غير موجودة في البيانات.")
        df["Needs Order Today"] = "Error"
        return df

    # تطبيق الشرط
    df["Needs Order Today"] = df.apply(
        lambda x: "Needs Order" if x["denom_id"] == "AED100" and x["Balance Today"] <= 100 else " ",
        axis=1
    )
    return df

# -----------------------------------------------------------
# def get_history(cashpoint_id):
#     try:
#         engine = get_db_engine()
#         query = '''
#             SELECT "Cashpoint ID", 'Date', 'Currency', 'Open. Bal', 'Deliveries',
#                 'Returns', 'Unpl. Del.', 'Unpl. Ret.', 'Deposits', 'Pre Service',
#                 'Withdrawals', 'Closing Bal.', 'H/E', 'Exclude'
#             FROM "History"
#             WHERE "Cashpoint ID" = :cashpoint_id
#             ORDER BY "Date" DESC
#         '''
#         df = pd.read_sql_query(text(query), con=engine, params={"cashpoint_id": cashpoint_id})
        
#         if df.empty:
#             return "No history available"

#         # تحويل DataFrame إلى سلسلة نصية
#         history_str = "\n".join(
#             f"{row['Date']}: {row['Action']} ({row['Component Type']})"
#             for _, row in df.iterrows()
#         )
#         return history_str

#     except Exception as e:
#         st.error(f"❌ Error fetching history for {cashpoint_id}: {str(e)}")
#         return "Error fetching history"

# def get_history(cashpoint_id):
#     try:
#         engine = get_db_engine()
#         query = '''
#             SELECT "Cashpoint ID", "Date", "Currency", "Open. Bal", "Deliveries",
#                         "Returns", "Unpl. Del.", "Unpl. Ret.", "Deposits", "Pre Service",
#                         "Withdrawals", "Closing Bal.", "H/E", "Exclude"
#             FROM "History"
#             WHERE "Cashpoint ID" = :cashpoint_id
#             ORDER BY "Date" DESC
#         '''
#         df = pd.read_sql_query(text(query), con=engine, params={"cashpoint_id": cashpoint_id})
        
#         if df.empty:
#             return "No history available"

#         # تحويل كل صف إلى سطر نصي منسّق
#         # history_str = "\n\n".join(
#         #     f"Date: {row['Date']}\n"
#         #     f"Currency: {row['Currency']}\n"
#         #     f"Open: {row['Open. Bal']}, Deliveries: {row['Deliveries']}, Returns: {row['Returns']}\n"
#         #     f"Unpl. Del.: {row['Unpl. Del.']}, Unpl. Ret.: {row['Unpl. Ret.']}\n"
#         #     f"Deposits: {row['Deposits']}, Pre Service: {row['Pre Service']}\n"
#         #     f"Withdrawals: {row['Withdrawals']}, Closing: {row['Closing Bal.']}\n"
#         #     f"H/E: {row['H/E']}, Exclude: {row['Exclude']}"
#         #     for _, row in df.iterrows()
#         # )

#         return df

#     except Exception as e:
#         st.error(f"❌ Error fetching history for {cashpoint_id}: {str(e)}")
#         return pd.DataFrame()

# 

# def get_history() -> pd.DataFrame:
#     try:
#         engine = get_db_engine()
#         query = '''
#             SELECT "Cashpoint ID", "Date", "Currency", "Open. Bal", "Deliveries",
#                     "Returns", "Unpl. Del.", "Unpl. Ret.", "Deposits", "Pre Service",
#                     "Withdrawals", "Closing Bal.", "H/E", "Exclude"
#             FROM "History"
#             ORDER BY "Date" DESC
#         '''
#         df = pd.read_sql_query(text(query), con=engine)

#         # الأعمدة النصية التي لا نريد تحويلها
#         string_columns = ["Cashpoint ID", "Date", "Currency"]

#         # الأعمدة الأخرى نحولها إلى Int64 (للسماح بالقيم الفارغة NaN)
#         for col in df.columns:
#             if col not in string_columns:
#                 # إزالة الفواصل إن وجدت (مثلاً 100,000 → 100000)
#                 df[col] = df[col].astype(str).str.replace(",", "")
#                 df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")

#         return df

#     except Exception as e:
#         st.error(f"❌ Error fetching full history: {str(e)}")
#         return pd.DataFrame()

def get_history() -> pd.DataFrame:
    try:
        engine = get_db_engine()
        query = '''
            SELECT *
            FROM "History"
            ORDER BY "Date" DESC
        '''
        df = pd.read_sql_query(text(query), con=engine)

        string_columns = ["Cashpoint ID", "Date", "Currency","H/E", "Exclude"]

        for col in df.columns:
            if col not in string_columns:
                df[col] = (
                    df[col]
                    .astype(str)
                    .str.replace(",", "", regex=False)
                    .str.strip()
                )
                df[col] = pd.to_numeric(df[col], errors="coerce")

        return df

    except Exception as e:
        st.error(f"❌ Error fetching full history: {str(e)}")
        return pd.DataFrame()


# -----------------------------------------------------------

def get_Forecast() -> pd.DataFrame:
    try:
        engine = get_db_engine()
        query = '''
            SELECT *
            FROM "Forecast"

        '''
                    # ORDER BY "date" DESC

        df = pd.read_sql_query(text(query), con=engine)

        # الأعمدة النصية التي لا نريد تحويلها
        # string_columns = []

        # الأعمدة الأخرى نحولها إلى Int64 (للسماح بالقيم الفارغة NaN)
        # for col in df.columns:
        #     if col not in string_columns:
        #         df[col] = (
        #             df[col]
        #             .astype(str)
        #             .str.replace(",", "", regex=False)
        #             .str.strip()
        #         )
        #         df[col] = pd.to_numeric(df[col], errors="coerce")

        return df

    except Exception as e:
        st.error(f"❌ Error fetching forecast data: {str(e)}")
        return pd.DataFrame()
    
# -----------------------------------------------------------

def get_parameters() -> pd.DataFrame:
    try:
        engine = get_db_engine()
        query = '''
            SELECT *

            FROM "Parameters"
        '''
        df = pd.read_sql_query(text(query), con=engine)
        
        # "Cashpoint ID", "Currency", "Minimum Delivery", "Min. Unplanned Delivery Amount", "Exception Amount",
        # "Exception % of Holdings", "Maximum Holding Amount", "Maximum Holding Type", "Optimization Level",
        # "Type", "Safety Stock", "Denom. Safety Stock", "% Adj. To Recommendations", "Pre-Deposit Percentage",
        # "Pre-Replenishment Percentage", "% Withdrawals Covered by Deposits", "Maximum Daily Recommendation",
        # "Standard Order Amount", "Insurance Costs", "Maximum Amount per Recommendation",
        # "ATM Add Cash Maximum Order Amount", "ATM Add Cash Maximum % Variance Recommended",
        # "Emergency ATM Add Cash Maximum Order Amount", "Emergency ATM Add Cash Maximum % Variance Recommended",
        # "ATM Replace Cash Maximum Order Amount", "ATM Replace Cash Maximum % Variance Recommended",
        # "Emergency ATM Replace Cash Maximum Order Amount", "Emergency ATM Replace Cash Maximum % Variance Recommended",
        # "ATM Return Cash Maximum Order Amount", "ATM Return Cash Maximum % Variance Recommended",
        # "Emergency ATM Return Maximum Order Amount", "Emergency ATM Return Maximum % Variance Recommended"

        # الأعمدة النصية التي لا نريد تحويلها
        # string_columns = []

        # الأعمدة الأخرى نحولها إلى Int64 (للسماح بالقيم الفارغة NaN)
        # for col in df.columns:
        #     if col not in string_columns:
        #         df[col] = (
        #             df[col]
        #             .astype(str)
        #             .str.replace(",", "", regex=False)
        #             .str.strip()
        #         )
        #         df[col] = pd.to_numeric(df[col], errors="coerce")

        return df

    except Exception as e:
        st.error(f"❌ Error fetching parameters data: {str(e)}")
        return pd.DataFrame()
    
# -----------------------------------------------------------

def get_Service() -> pd.DataFrame:
    try:
        engine = get_db_engine()
        query = '''
            SELECT *
            FROM "Service Day"

        '''
        
            # "OptiCash - Cashpoint Service Days Report - ATM",
            # "Unnamed: 1", "Unnamed: 2", "Unnamed: 3", "Unnamed: 4",
            # "Transit Time", "Unnamed: 6",
            # "Add Cash Days", "Unnamed: 8", "Unnamed: 9", "Unnamed: 10", "Unnamed: 11",
            # "Replace Cash Days", "Unnamed: 13", "Unnamed: 14", "Unnamed: 15", "Unnamed: 16",
            # "Unplanned Service", "Unnamed: 18",
            # "ATM Return Days", "Unnamed: 20", "Unnamed: 21", "Unnamed: 22", "Unnamed: 23"
            
            
        df = pd.read_sql_query(text(query), con=engine)
        
        # ✅ إعادة تسمية الأعمدة يدويًا
        new_columns = [
            "ATM", "Unnamed: 1", "Unnamed: 2", "Unnamed: 3", "Unnamed: 4",
            "Transit Time", "Unnamed: 6",
            "Add Cash Days", "Unnamed: 8", "Unnamed: 9", "Unnamed: 10", "Unnamed: 11",
            "Replace Cash Days", "Unnamed: 13", "Unnamed: 14", "Unnamed: 15", "Unnamed: 16",
            "Unplanned Service", "Unnamed: 18",
            "ATM Return Days", "Unnamed: 20", "Unnamed: 21", "Unnamed: 22", "Unnamed: 23"
        ]

        if len(df.columns) == len(new_columns):
            df.columns = new_columns
        else:
            st.warning("⚠️ عدد الأعمدة في الجدول لا يطابق عدد الأسماء المحددة!")
        
        # تجاهل أول صف
        # df = df.iloc[1:].reset_index(drop=True)
        
        # ✅ تحويل الصف الأول إلى أسماء أعمدة
        # df.columns = df.iloc[0]  # الصف الأول يصبح رؤوس الأعمدة
        # df = df[1:].reset_index(drop=True)  # حذف الصف الأول الفعلي وإعادة الفهرسة

        # الأعمدة النصية التي لا نريد تحويلها
        # string_columns = []

        # الأعمدة الأخرى نحولها إلى Int64 (للسماح بالقيم الفارغة NaN)
        # for col in df.columns:
        #     if col not in string_columns:
        #         df[col] = (
        #             df[col]
        #             .astype(str)
        #             .str.replace(",", "", regex=False)
        #             .str.strip()
        #         )
        #         df[col] = pd.to_numeric(df[col], errors="coerce")

        return df

    except Exception as e:
        st.error(f"❌ Error fetching service day data: {str(e)}")
        return pd.DataFrame()

# -----------------------------------------------------------

def get_DFF():
    try:
        engine = get_db_engine()
        query = '''
            SELECT *
            FROM "DFF"
        '''
        df = pd.read_sql_query(text(query), con=engine)

        # الأعمدة النصية التي لا نريد تحويلها
        # string_columns = []

        # الأعمدة الأخرى نحولها إلى Int64 (للسماح بالقيم الفارغة NaN)
        # for col in df.columns:
        #     if col not in string_columns:
        #         df[col] = (
        #             df[col]
        #             .astype(str)
        #             .str.replace(",", "", regex=False)
        #             .str.strip()
        #         )
        #         df[col] = pd.to_numeric(df[col], errors="coerce")

        return df

    except Exception as e:
        st.error(f"❌ Error fetching DFF data: {str(e)}")
        return pd.DataFrame()
    
# -----------------------------------------------------------

def get_enhanced():
    try:
        engine = get_db_engine()
        query = '''
            SELECT *
            FROM "Enhanced"
        '''
        df = pd.read_sql_query(text(query), con=engine)

        # الأعمدة النصية التي لا نريد تحويلها
        # string_columns = []

        # الأعمدة الأخرى نحولها إلى Int64 (للسماح بالقيم الفارغة NaN)
        # for col in df.columns:
        #     if col not in string_columns:
        #         df[col] = (
        #             df[col]
        #             .astype(str)
        #             .str.replace(",", "", regex=False)
        #             .str.strip()
        #         )
        #         df[col] = pd.to_numeric(df[col], errors="coerce")

        return df

    except Exception as e:
        st.error(f"❌ Error fetching enhanced data: {str(e)}")
        return pd.DataFrame()

# -----------------------------------------------------------

def get_horizon():
    try:
        engine = get_db_engine()
        query = '''
            SELECT *
            FROM "Horizon"
        '''
        df = pd.read_sql_query(text(query), con=engine)

        # الأعمدة النصية التي لا نريد تحويلها
        # string_columns = []

        # الأعمدة الأخرى نحولها إلى Int64 (للسماح بالقيم الفارغة NaN)
        # for col in df.columns:
        #     if col not in string_columns:
        #         df[col] = (
        #             df[col]
        #             .astype(str)
        #             .str.replace(",", "", regex=False)
        #             .str.strip()
        #         )
        #         df[col] = pd.to_numeric(df[col], errors="coerce")

        return df

    except Exception as e:
        st.error(f"❌ Error fetching horizon data: {str(e)}")
        return pd.DataFrame()

# -----------------------------------------------------------

def get_horizon_advanced():
    try:
        engine = get_db_engine()
        query = '''
            SELECT *
            FROM "Advanced Horizon"
        '''
        df = pd.read_sql_query(text(query), con=engine)

        # الأعمدة النصية التي لا نريد تحويلها
        # string_columns = []

        # الأعمدة الأخرى نحولها إلى Int64 (للسماح بالقيم الفارغة NaN)
        # for col in df.columns:
        #     if col not in string_columns:
        #         df[col] = (
        #             df[col]
        #             .astype(str)
        #             .str.replace(",", "", regex=False)
        #             .str.strip()
        #         )
        #         df[col] = pd.to_numeric(df[col], errors="coerce")

        return df

    except Exception as e:
        st.error(f"❌ Error fetching advanced horizon data: {str(e)}")
        return pd.DataFrame()

# -----------------------------------------------------------

def get_Recommendation():
    try:
        engine = get_db_engine()
        query = '''
            SELECT *
            FROM "Recommendation"
        '''
        df = pd.read_sql_query(text(query), con=engine)

        # الأعمدة النصية التي لا نريد تحويلها
        # string_columns = []

        # الأعمدة الأخرى نحولها إلى Int64 (للسماح بالقيم الفارغة NaN)
        # for col in df.columns:
        #     if col not in string_columns:
        #         df[col] = (
        #             df[col]
        #             .astype(str)
        #             .str.replace(",", "", regex=False)
        #             .str.strip()
        #         )
        #         df[col] = pd.to_numeric(df[col], errors="coerce")

        return df

    except Exception as e:
        st.error(f"❌ Error fetching recommendation data: {str(e)}")
        return pd.DataFrame()

# -----------------------------------------------------------
# Get Horizon Data
# def get_horizon_data(df):

#     try:
#         engine = get_db_engine()
#         horizon_query = 'SELECT DISTINCT "Cashpoint ID" FROM "Horizon"'
#         advanced_horizon_query = 'SELECT DISTINCT "Cashpoint ID" FROM "Advanced Horizon"'
#         horizon_df = pd.read_sql_query(text(horizon_query), con=engine)
#         advanced_horizon_df = pd.read_sql_query(text(advanced_horizon_query), con=engine)

#         # if horizon_df.empty:
#         #     st.warning("جدول 'Horizon' فارغ.")
#         #     df["Horizon"] = " "
#         #     return df

#         # قائمة الماكينات الموجودة في جدول Horizon
#         horizon_ids = set(horizon_df["Cashpoint ID"].dropna().unique())
#         advanced_horizon_ids = set(advanced_horizon_df["Cashpoint ID"].dropna().unique())

#         # ✅ إنشاء العمود الجديد بنفس طول df
#         df["Horizon"] = df["cashp_id"].apply(
#             lambda x: x if x in horizon_ids else " "
#         )
#         df["Advanced Horizon"] = df["cashp_id"].apply(
#             lambda x: x if x in advanced_horizon_ids else " "
#         )   

#         return df

#     except Exception as e:
#         st.error(f"❌ Error fetching horizon data: {str(e)}")
#         df["Horizon"] = "Error"
#         return df

# def get_horizon_data():
#     try:
#         engine = get_db_engine()
        
#         # جلب الماكينات من جداول Horizon و Advanced Horizon
#         horizon_query = 'SELECT DISTINCT "Cashpoint ID" FROM "Horizon"'
#         advanced_horizon_query = 'SELECT DISTINCT "Cashpoint ID" FROM "Advanced Horizon"'

#         horizon_df = pd.read_sql_query(text(horizon_query), con=engine)
#         advanced_horizon_df = pd.read_sql_query(text(advanced_horizon_query), con=engine)

#         horizon_ids = horizon_df["Cashpoint ID"]
#         advanced_horizon_ids = advanced_horizon_df["Cashpoint ID"]

#         # دمج القائمتين
#         total_ids = horizon_ids.union(advanced_horizon_ids)

#         # فلترة df الأصلي بناءً على الـ IDs الموجودة في الجداول
#         # df_filtered = df[df["cashp_id"].isin(total_ids)].copy()

#         # إضافة عمودين للمعلومية فقط
#         # df_filtered["Horizon"] = df_filtered["cashp_id"].apply(
#         #     lambda x: x if x in horizon_ids else " "
#         # )
#         # df_filtered["Advanced Horizon"] = df_filtered["cashp_id"].apply(
#         #     lambda x: x if x in advanced_horizon_ids else " "
#         # )

#         return total_ids

#     except Exception as e:
#         st.error(f"❌ Error fetching horizon data: {str(e)}")
#         return pd.DataFrame()

# def get_horizon_data():
#     try:
#         engine = get_db_engine()
        
#         # جلب الماكينات من جداول Horizon و Advanced Horizon
#         horizon_query = 'SELECT DISTINCT "Cashpoint ID", "Date", "Open. Bal", "Withdrawals", "Closing Bal." FROM "Horizon"'
#         advanced_horizon_query = 'SELECT DISTINCT "Cashpoint ID", "Date", "Open Bal", "Withdrawals", "Clos Bal" FROM "Advanced Horizon"'

#         horizon_df = pd.read_sql_query(text(horizon_query), con=engine)
#         advanced_horizon_df = pd.read_sql_query(text(advanced_horizon_query), con=engine)

#         # # ✅ تحويل Series إلى set قبل union
#         # horizon_ids = set(horizon_df["Cashpoint ID","Date","Open. Bal","Withdrawals","Closing Bal."])
#         # advanced_horizon_ids = set(advanced_horizon_df["Cashpoint ID","Date","Open Bal","Withdrawals","Clos Bal"])
        
#         horizon_ids = set(tuple(row) for row in horizon_df[["Cashpoint ID", "Date", "Open. Bal", "Withdrawals", "Closing Bal."]].values)
#         advanced_horizon_ids = set(tuple(row) for row in advanced_horizon_df[["Cashpoint ID", "Date", "Open Bal", "Withdrawals", "Clos Bal"]].values)

        
#         def is_valid(val):
#             val = str(val).lower().strip()
#             return val not in ["summary", "average", "total"]

#         # filtered_ids = list(filter(is_valid, horizon_ids.union(advanced_horizon_ids)),asending=True)
#         filtered_ids = sorted(filter(is_valid, horizon_ids.union(advanced_horizon_ids)))
#         return pd.DataFrame({"Cashpoint ID,Date,": filtered_ids})

#         # # دمج القائمتين بدون تكرار
#         # total_ids = horizon_ids.union(advanced_horizon_ids)

#         # # ✅ إرجاعهم كـ DataFrame
#         # return pd.DataFrame({"Cashpoint ID": list(total_ids)})

#     except Exception as e:
#         st.error(f"❌ Error fetching horizon data: {str(e)}")
#         return pd.DataFrame()

# def get_horizon_data():
#     try:
#         engine = get_db_engine()
        
#         # استعلامات SQL
#         horizon_query = '''
#             SELECT DISTINCT "Cashpoint ID", "Date", "Open. Bal", "Withdrawals", "Closing Bal." 
#             FROM "Horizon"
#         '''
#         advanced_horizon_query = '''
#             SELECT DISTINCT "Cashpoint ID", "Date", "Open Bal", "Withdrawals", "Clos Bal", "Component Type" 
#             FROM "Advanced Horizon"
#         '''

#         # قراءة البيانات
#         horizon_df = pd.read_sql_query(text(horizon_query), con=engine)
#         advanced_horizon_df = pd.read_sql_query(text(advanced_horizon_query), con=engine)

#         # إعادة تسمية الأعمدة لتوحيدها
#         horizon_df = horizon_df.rename(columns={
#             "Open. Bal": "Open Bal",
#             "Closing Bal.": "Clos Bal"
#         })
#         advanced_horizon_df = advanced_horizon_df.rename(columns={
#             "Open Bal": "Open Bal",
#             "Clos Bal": "Clos Bal"
#         })

#         # فلترة فقط Withdrawals Only
#         if "Component Type" in advanced_horizon_df.columns:
#             advanced_horizon_df = advanced_horizon_df[advanced_horizon_df["Component Type"] == "Withdrawals Only"]
#             advanced_horizon_df = advanced_horizon_df.drop(columns=["Component Type"])

#         # تنظيف التاريخ من الرموز مثل S / R / U
#         # def clean_date_column(df):
#         #     df["Date"] = df["Date"].astype(str).str.extract(r'(\d{4}-\d{2}-\d{2})')[0]
#         #     df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
#         #     return df
        
#         # def clean_date_column(df):
#         #     df["Date"] = df["Date"].astype(str).str.extract(r'(\d{4}-\d{2}-\d{2})') #[0]
#         #     df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date  # ⬅️ إزالة الوقت
#         #     return df
        
#         def clean_date_column(df):
#             def parse_date(val):
#                 val = str(val)
#                 # استخراج أول تطابق لتاريخ بأي من الصيغتين
#                 match = re.search(r'(\d{4}-\d{2}-\d{2})|(\d{2}-\d{2}-\d{4})', val)
#                 if match:
#                     date_str = match.group(0)
#                     try:
#                         # نجرب yyyy-mm-dd
#                         return pd.to_datetime(date_str, format="%Y-%m-%d").date()
#                     except:
#                         try:
#                             # نجرب dd-mm-yyyy
#                             return pd.to_datetime(date_str, dayfirst=True).date()
#                         except:
#                             return None
#                 return None

#             df["Date"] = df["Date"].apply(parse_date)
#             return df

#         horizon_df = clean_date_column(horizon_df)
#         advanced_horizon_df = clean_date_column(advanced_horizon_df)

#         # دمج الجدولين
#         combined_df = pd.concat([horizon_df, advanced_horizon_df], ignore_index=True)

#         # حذف الصفوف غير المرغوبة مثل Summary / Average / Total
#         combined_df = combined_df[~combined_df["Cashpoint ID"].astype(str).str.lower().isin(["summary", "average", "total"])]

#         # حذف الصفوف التي تحتوي على بيانات ناقصة
#         # combined_df = combined_df.dropna()

#         # إعادة الترتيب حسب Cashpoint ID وDate
#         combined_df = combined_df.sort_values(by=["Date","Cashpoint ID", "Date"], ascending=True)

#         return combined_df

#     except Exception as e:
#         st.error(f"❌ Error fetching horizon data: {str(e)}")
#         return pd.DataFrame()

def get_horizon_data():
    try:
        engine = get_db_engine()

        # استعلامات SQL
        horizon_query = '''
            SELECT DISTINCT "Cashpoint ID", "Date", "Open. Bal", "Withdrawals", "Closing Bal." 
            FROM "Horizon"
        '''
        advanced_horizon_query = '''
            SELECT DISTINCT "Cashpoint ID", "Date", "Open Bal", "Withdrawals", "Clos Bal", "Component Type" 
            FROM "Advanced Horizon"
        '''

        # قراءة البيانات
        horizon_df = pd.read_sql_query(text(horizon_query), con=engine)
        advanced_horizon_df = pd.read_sql_query(text(advanced_horizon_query), con=engine)

        # إعادة تسمية الأعمدة لتوحيدها
        horizon_df = horizon_df.rename(columns={
            "Open. Bal": "Open Bal",
            "Closing Bal.": "Clos Bal"
        })
        advanced_horizon_df = advanced_horizon_df.rename(columns={
            "Open Bal": "Open Bal",
            "Clos Bal": "Clos Bal"
        })

        # فلترة Withdrawals Only من Advanced فقط
        advanced_horizon_df = advanced_horizon_df[advanced_horizon_df["Component Type"] == "Withdrawals Only"]
        advanced_horizon_df = advanced_horizon_df.drop(columns=["Component Type"])

        # تنظيف التاريخ باستخدام dateutil
        def clean_date_column(df):
            def try_parse(val):
                try:
                    return parse(str(val), dayfirst=True, fuzzy=True).date()
                except:
                    return None
            df["Date"] = df["Date"].apply(try_parse)
            return df

        horizon_df = clean_date_column(horizon_df)
        advanced_horizon_df = clean_date_column(advanced_horizon_df)

        # دمج البيانات
        combined_df = pd.concat([horizon_df, advanced_horizon_df], ignore_index=True)

        # حذف الصفوف غير المرغوبة
        combined_df = combined_df[~combined_df["Cashpoint ID"].astype(str).str.lower().isin(["summary", "average", "total"])]

        # ترتيب حسب التاريخ والماكينة
        combined_df = combined_df.sort_values(by=["Date", "Cashpoint ID"])
        
        today = date.today()
        # today = date(2024,8,25)
        open_today_df = combined_df[combined_df["Date"] == today][["Cashpoint ID", "Open Bal","Withdrawals"]].copy()
        open_today_df = open_today_df.rename(columns={"Withdrawals": "Forecast"})

        return combined_df, open_today_df

    except Exception as e:
        st.error(f"❌ Error fetching horizon data: {str(e)}")
        return pd.DataFrame()

# -----------------------------------------------------------

#Get Open Horizon of today 
    # استخراج رصيد الفتح لماكينات اليوم
    


# -----------------------------------------------------------
# # Get Enhanced ATMs
# def get_check_recommendation(df):
#     try:
#         engine = get_db_engine()
#         enhanced_query = 'SELECT DISTINCT "Cashpoint ID", "Component Type", "Withdrawals", "Closing Bal." FROM "Enhanced"'
#         enhanced_df = pd.read_sql_query(text(enhanced_query), con=engine)

#         # if enhanced_df.empty:
#         #     st.warning("جدول 'Enhanced ATMs' فارغ.")
#         #     df["Enhanced ATMs"] = " "
#         #     return df
        

#         Withdrawals_Only = enhanced_df["Component Type"] == "Withdrawals Only"
        
#         # إذا كان هناك شرط إضافي، يمكنك تطبيقه هنا
#         sum_withdrawals = enhanced_df[Withdrawals_Only].groupby("Cashpoint ID")["Withdrawals"].sum().reset_index()
#         sum_closing_balance = enhanced_df[~Withdrawals_Only].groupby("Cashpoint ID")["Closing Bal."].sum().reset_index()
        
#         # قائمة الماكينات الموجودة في جدول Enhanced ATMs
#         enhanced_ids = set(enhanced_df["Cashpoint ID"])

#         # # ✅ إنشاء العمود الجديد بنفس طول df
#         df["Enhanced ATMs"] = df["Cashpoint ID"].apply(
#             lambda x: x if x in enhanced_ids else " "
#         )

#         return df

#     except Exception as e:
#         st.error(f"❌ Error fetching enhanced ATMs: {str(e)}")
#         df["Enhanced ATMs"] = "Error"
#         return df


def get_check_recommendation(df, open_today_df):
    try:
        engine = get_db_engine()
        query = '''
            SELECT "Cashpoint ID", "Component Type", "Withdrawals", "Closing Bal."
            FROM "Enhanced"
        '''
        df = pd.read_sql_query(text(query), con=engine)

        # تصفية البيانات على Withdrawals Only
        withdrawals_df = df[df["Component Type"] == "Withdrawals Only"]

        # تجميع المبالغ
        summary = withdrawals_df.groupby("Cashpoint ID", as_index=False).agg({
            "Withdrawals": "sum",
            "Closing Bal.": "sum"
        })

        # إعادة تسمية الأعمدة للوضوح
        summary = summary.rename(columns={
            "Withdrawals": "Sum_Withdrawals",
            "Closing Bal.": "Sum_Closing_balance",
        })
        
        # summary["Cashpoint ID"] = summary["Cashpoint ID"].astype(str).str.strip()
        # open_today_df["Cashpoint ID"] = open_today_df["Cashpoint ID"].astype(str).str.strip()
        
        # دمج open_today_df بناءً على Cashpoint ID
        summary = summary.merge(open_today_df, on="Cashpoint ID", how="left")
        # إنشاء عمود اختبار للمقارنة بين العمودين
        summary["Test Recom."] = (summary["Sum_Closing_balance"].astype("Int64") == summary["Open Bal"].astype("Int64")).astype(bool)
        
        #أنشاء عمود الاختلاف بين ال closing balance و Forecast
        summary["Difference Forecast"] = summary["Sum_Closing_balance"] - summary["Forecast"]
        
        # إدراج آخر Withdrawals من الهيستوري
        # summary["History"] = summary["Cashpoint ID"].apply(
        #     lambda x: get_history(x) if x in df["Cashpoint ID"].values else pd.DataFrame()
        # )
        # summary["History"] = summary["Cashpoint ID"].apply(
        #     lambda x: (
        #         history := get_history(),
        #         int(str(history["Withdrawals"].iloc[0]).replace(",", "")) 
        #         if not history.empty and "Withdrawals" in history.columns else None)[1]
        #     ).astype("Int64")
        
        # # إدراج آخر Withdrawals من الهيستوري
        # summary["History"] = summary["Cashpoint ID"].apply(
        #     lambda x: get_history()["Withdrawals"].iloc[0]
        #     if isinstance(get_history(), pd.DataFrame)
        #     and not get_history().empty
        #     and "Withdrawals" in get_history().columns
        #     else None
        # ).astype("Int64")
# ✅ جلب بيانات History واستخراج آخر Withdrawals لكل ماكينة
        df_history = get_history()

        if not df_history.empty:
            df_history["Date"] = pd.to_datetime(df_history["Date"], errors="coerce")
            df_history_sorted = df_history.sort_values(by=["Cashpoint ID", "Date"], ascending=[True, False])
            df_history_latest = df_history_sorted.drop_duplicates(subset=["Cashpoint ID"], keep="first")

            summary = summary.merge(
                df_history_latest[["Cashpoint ID", "Withdrawals"]],
                on="Cashpoint ID",
                how="left"
            )
            summary = summary.rename(columns={"Withdrawals": "History"})
            summary["History"] = summary["History"].astype("Int64")
        else:
            summary["History"] = pd.NA
        # إنشاء عمود اختبار للمقارنة بين Sum_Withdrawals و History
        summary["Test History"] = (summary["Sum_Withdrawals"].astype("Int64") == summary["History"].astype("Int64")
            ).astype(bool)

        summary["Difference History"] = summary["Sum_Closing_balance"] - summary["History"]
        
        # # أو تحقق كلي:
        # if (summary["sum_closing_balance"] == summary["Open Bal"]).all():
        #     print("كل القيم متساوية")
        # else:
        #     print("بعض القيم مختلفة")

        
        # إعادة ترتيب الأعمدة: (Cashpoint ID, sum_withdrawals, sum_closing_balance, Open Bal, Withdrawals)
        summary = summary[["Cashpoint ID", "Sum_Withdrawals", "Sum_Closing_balance", "Open Bal", "Forecast", "Test Recom.",
                            "Difference Forecast", "History", "Test History", "Difference History"]]

        # عرض النتيجة
        # st.dataframe(summary)
        
        return summary

    except Exception as e:
        st.error(f"❌ Error generating withdrawals summary: {str(e)}")


# -----------------------------------------------------------

# Get Recommendation Orders
def get_recommendation_orders(df):
    try:
        engine = get_db_engine()
        recomm_query = 'SELECT DISTINCT "Cashpoint ID" FROM "Recommendation"'
        recomm_df = pd.read_sql_query(text(recomm_query), con=engine)
        # emergency_query = "SELECT DISTINCT \"Cashpoint ID\" FROM \"Recommendation\" WHERE \"Action\" = 'Emergency ATM Replace Cash'"
        # emerg_df = pd.read_sql_query(text(emergency_query), con=engine)

        # if recomm_df.empty:
        #     st.warning("جدول 'Recommendation' فارغ.")
        #     df["Recomm. Orders"] = " "
        #     # df["Emerg. Orders"] = " "
        #     return df

        # قائمة الماكينات الموجودة في جدول Recommendation
        recommended_ids = set(recomm_df["Cashpoint ID"].dropna().unique())
        # Emergency_ids = set(emerg_df["Cashpoint ID"].dropna().unique())

        # ✅ إنشاء العمود الجديد بنفس طول df
        df["Recomm. Orders"] = df["cashp_id"].apply(
            lambda x: x if x in recommended_ids else " "
        )
    
        # # ✅إنشاء العمود الجديد بنفس طول df
        # df["Emerg. Orders"] = df["cashp_id"].apply(
        #     lambda x: x if x in Emergency_ids else " "
        # )

        return df

    except Exception as e:
        st.error(f"❌ Error fetching recommendation orders: {str(e)}")
        df["Recomm. Orders"] = "Error"
        return df
# -----------------------------------------------------------

# Get Emergency Orders
def get_emergency_orders(df):
    try:
        engine = get_db_engine()
        emergency_query ="SELECT DISTINCT \"Cashpoint ID\" FROM \"Recommendation\" WHERE \"Action\" = 'Emergency ATM Replace Cash'"
        emerg_df = pd.read_sql_query(text(emergency_query), con=engine)

        # if emerg_df.empty:
        #     st.warning("جدول 'Recommendation' فارغ.")
        #     df["Emerg. Orders"] = " "
        #     return df

        # قائمة الماكينات الموجودة في جدول Recommendation
        Emergency_ids = set(emerg_df["Cashpoint ID"].dropna().unique())

        # ✅ إنشاء العمود الجديد بنفس طول df
        df["Emerg. Orders"] = df["cashp_id"].apply(
            lambda x: x if x in Emergency_ids else " "
        )

        return df

    except Exception as e:
        st.error(f"❌ Error fetching emergency orders: {str(e)}")
        df["Emerg. Orders"] = "Error"
        return df

# -----------------------------------------------------------

# Get Orders in Plan
def get_orders_in_plan(df):
    try:
        engine = get_db_engine()
        plan_query = 'SELECT DISTINCT "Cashpoint ID" FROM "Plan"'
        plan_df = pd.read_sql_query(text(plan_query), con=engine)

        # if plan_df.empty:
        #     st.warning("جدول 'Plan' فارغ.")
        #     df["Plan Orders"] = " "
        #     return df

        # قائمة الماكينات الموجودة في جدول Plan
        plan_ids = set(plan_df["Cashpoint ID"].dropna().unique())

        # ✅ إنشاء العمود الجديد بنفس طول df
        df["Plan Orders"] = df["cashp_id"].apply(
            lambda x: x if x in plan_ids else " "
        )

        return df

    except Exception as e:
        st.error(f"❌ Error fetching orders in plan: {str(e)}")
        df["Plan Orders"] = "Error"
        return df

# -----------------------------------------------------------
# دالة جلب الماكينات التي صرفت أكثر من 100,000
def get_high_dispense_atms_ids():
    try:
        engine = get_db_engine()
        query = """
            SELECT
                Forecast."Cashpoint ID",
                Forecast."Date",
                Forecast."Withdrawal-Actual",
                Forecast."Withdrawal-Forecast"
            FROM "Forecast" AS Forecast
        """
        df = pd.read_sql_query(text(query), con=engine)
        df["withdrawal_diff"] = df["Withdrawal-Actual"] - df["Withdrawal-Forecast"]
        high_diss = df[df["withdrawal_diff"] > 100000]
        return high_diss
    except Exception as e:
        st.error(f"❌ Error fetching high dispense ATMs: {str(e)}")
        return []


# # Get High Dispense ATMs
# def get_high_dispense_atms(df):
#     try:
#         engine = get_db_engine()
#         high_dispense_query = """
#         SELECT
#             Forecast."cashpoint ID",
#             Forecast."Date",
#             Forecast."Withdrawal_Actual",
#             Forecast."Withdrawal_Forecast"
        
#         FROM "Forecast" as Forecast
        
#         """
#         df = pd.read_sql_query(text(high_dispense_query), con=engine)
        
#         df["withdrawal_diff"] = df["Withdrawal-Actual"] - df["Withdrawal-Forecast"]
        
#         # فلترة القيم الأكبر من 100,000
#         df_filtered_high_dispense = df[df["withdrawal_diff"] > 100000]

#         if df_filtered_high_dispense.empty:
#             st.warning("جدول 'High Dispense ATMs' فارغ.")
#             df["High Dispense ATMs"] = " "
#             return df

#         # قائمة الماكينات الموجودة في جدول High Dispense ATMs
#         high_dispense_ids = set(df_filtered_high_dispense["cashpoint_ID"].dropna().unique())

#         # ✅ إنشاء العمود الجديد بنفس طول df
#         df["High Dispense ATMs"] = df["cashpoint_ID"].apply(
#             lambda x: x if x in high_dispense_ids else " "
#             )

#         return df_filtered_high_dispense

#     except Exception as e:
#         st.error(f"❌ Error fetching high dispense ATMs: {str(e)}")
#         df["High Dispense ATMs"] = "Error"
#         return df


# Get list of High Dispense ATM IDs only
# def get_high_dispense_atms():
#     try:
#         engine = get_db_engine()
#         query = """
#         SELECT
#             Forecast."CashpointID",
#             Forecast."Date",
#             Forecast."Withdrawal-Actual",
#             Forecast."Withdrawal-Forecast"
#         FROM "Forecast" as Forecast
#         """
#         df = pd.read_sql_query(text(query), con=engine)
#         df["withdrawal_diff"] = df["Withdrawal-Actual"] - df["Withdrawal-Forecast"]
#         df_filtered = df[df["withdrawal_diff"] > 100000]

#         return df_filtered["cashpoint_ID"].dropna().unique()

#     except Exception as e:
#         st.error(f"❌ Error fetching high dispense ATMs: {str(e)}")
#         return []


# -----------------------------------------------------------
high_diss_ids = get_high_dispense_atms_ids()

# -----------------------------------------------------------
# Get High Dispense ATMs
def get_order_in_plan(df):
    try:
        engine = get_db_engine()
        plan_query = """
        SELECT
            "Due Date",
            "Cashpoint ID",
            "CashPoint Name",
            "Cashpoint Type",
            "Action",
            "Conf. #",
            "Order Date",
            "Order Src.",
            "Override Reason",
            "Currency",
            "Amount",
            "Depot ID"
        FROM "Plan"
        """
        df = pd.read_sql_query(text(plan_query), con=engine)

        if df.empty:
            st.warning("جدول 'Plan Orders' فارغ.")
            df["Plan Orders"] = " "
            return df
        
        # Plan_ids = set(plan_df["Cashpoint ID"].dropna().unique())

        # # ✅ إنشاء العمود الجديد بنفس طول df
        # df["Plan Orders"] = df["Cashpoint ID"].apply(
        #     lambda x: x if x in Plan_ids else " "
        # )

        return df.sort_values(by=["Due Date","Cashpoint ID"], ascending=True)

    except Exception as e:
        st.error(f"❌ Error fetching plan orders: {str(e)}")
        df["Plan Orders"] = "Error"
        return df

# -----------------------------------------------------------

# 3- Get Common DFF-ATMs Data
# @st.cache_data(ttl=600)
def get_common_names_between_ATMs_and_DFF():
    try:
        engine = get_db_engine()
        query = """
        SELECT 
            dff.cashp_id,
            dff.denom_id,
            dff.open_bal,
            dff.nopen_bal,
            dff.norm_del,
            dff.nnorm_del,
            dff.norm_rtr,
            dff.nnorm_rtr,
            dff.wthdrwls,
            dff.nwthdrwls,
            dff.clos_bal,
            dff.nclos_bal
        FROM "DFF" dff
        INNER JOIN "ATMs" atm ON dff.cashp_id = atm."Cashpoint ID"
        """
        df = pd.read_sql_query(text(query), con=engine)

        # if df.empty:
        #     st.warning("No data found in DFF and ATMs tables.")
        #     return pd.DataFrame()

        df["ATMs"] = df["cashp_id"]
        # df["Cass. Cover Days"] = conve cassette_cover_days(df)
        # # df["Cass. Cover Days"] = cover_nan_to_str(df["Cass. Cover Days"])
        # df["Balance Today"] = Balance_Today(df)["Balance Today"],cover_nan_to_str(df["Balance Today"])
        # df["Balance Tomorrow"] = Balance_Tomorrow(df)["Balance Tomorrow"], cover_nan_to_str(df["Balance Tomorrow"])
        # # df["Balance Tomorrow"] = cover_nan_to_str(df["Balance Tomorrow"])
        # # df["Balance Today"] = cover_nan_to_str(df["Balance Today"])
        
        # df["Cass. Cover Days"] = cassette_cover_days(df)
        # df = Balance_Today(df)
        # df = Balance_Tomorrow(df)

        # بعد العمليات الحسابية، طبق التحويل إلى نص مرة واحدة لكل عمود
        df["Cass. Cover Days"] = cassette_cover_days(df)
        df["Cass. Cover Days"] = cover_nan_to_str(df["Cass. Cover Days"])
        df["Balance Today"] = Balance_Today(df)["Balance Today"]
        df["Balance Tomorrow"] = Balance_Tomorrow(df)["Balance Tomorrow"]
        # df["Balance Today"] = cover_nan_to_str(df["Balance Today"])
        # df["Balance Tomorrow"] = cover_nan_to_str(df["Balance Tomorrow"])
        df["Recomm. Orders"] = get_recommendation_orders(df)["Recomm. Orders"]
        df["Emerg. Orders"] = get_emergency_orders(df)["Emerg. Orders"]  # تعيين القيمة الافتراضية
        df["Plan Orders"] = get_orders_in_plan(df)["Plan Orders"]  # تعيين القيمة الافتراضية
        df["Needs Order Today"] = get_needs_order_today(df)["Needs Order Today"]  # تعيين القيمة الافتراضية
        df["High Dispense ATMs"] = df["cashp_id"].where(df["cashp_id"].isin(high_diss_ids), "")
        
        return df

    except Exception as e:
        st.error(f"❌ Error: {str(e)}")
        return pd.DataFrame()

#&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&

# -----------------------------------------------------------
# 4- Convert DataFrame to Excel


# def apply_red_font_to_negatives(worksheet, column_name="Balance Today, Balance Tomorrow"):
#     for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
#         for cell in row:
#             if isinstance(cell.value, str) and cell.value.startswith('(') and cell.value.endswith(')'):
#                 cell.font = Font(color='FF0000', bold=True, size=8)

# def apply_red_font_for_negatives(worksheet):
#     for row in worksheet.iter_rows(min_row=2):  # تجاهل رأس الجدول
#         for cell in row:
#             if isinstance(cell.value, (int, float)) and cell.value < 0:
#                 # cell.value = f"({abs(cell.value)})"# تحويل القيمة السالبة إلى تنسيق بين قوسين 
#                 cell.number_format = '#,##0;(#,##0)'   # يعرض الرقم بين أقواس دون تغيير القيمة
#                 cell.font = Font(color="FF0000", bold=True, size=8)  # تطبيق خط أحمر عريض

def apply_red_font_for_negatives(worksheet):
    for row in worksheet.iter_rows(min_row=2):  # تجاهل رأس الجدول
        for cell in row:
            try:
                val = float(cell.value)
                if val < 0:
                    cell.number_format = '#,##0;(#,##0)'
                    cell.font = Font(color="FF0000", bold=True, size=8)
                else:
                    cell.number_format = '#,##0'
            except (TypeError, ValueError):
                # الخلية ليست قيمة رقمية، تجاهلها
                pass


@st.cache_data
def convert_to_excel(data, sheet_name="Sheet1"):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    data.to_excel(writer, sheet_name=sheet_name, index=False)
    workbook = writer.book
    worksheet = workbook[sheet_name]
    # default_font = Font(size=8)

# تنسيقات الخط والحدود والألوان
def style_worksheet(worksheet):
    bold_font = Font(bold=True, size=8)
    white_font = Font(color='FFFFFF', bold=True, size=8)
    blue_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    border = Border(left=Side(style='double', color='0000FF'),
                    right=Side(style='double', color='0000FF'),
                    top=Side(style='double', color='0000FF'),
                    bottom=Side(style='double', color='0000FF'))

# تنسيق البيانات (بدءًا من الصف الثاني، لأن الصف الأول رؤوس)
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.font = bold_font
            cell.border = border
            cell.alignment =Alignment(horizontal='center', vertical='center')
            
        # ✅ تنسيق القيم الحقيقية
        # if row[1].value == "AED100":
        if len(row) > 1 and row[1].value == "AED100":
            for cell in row[1:14]:
                cell.fill = orange_fill

# تنسيق رؤوس الأعمدة (الصف الأول)
    for cell in worksheet[1]:
        cell.font = white_font

        cell.fill = blue_fill
        cell.border = border
        cell.alignment = cell.alignment.copy(horizontal='center', vertical='center')

# ضبط عرض الأعمدة تلقائيًا
    for col_idx, col_cells in enumerate(worksheet.columns, 1):
        max_length = max((len(str(cell.value)) for cell in col_cells if cell.value), default=0)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length +1


# 1. تجميد الصف الأول (Freeze Pane)
    worksheet.freeze_panes = "A2"

# 2. تفعيل الفلتر (AutoFilter) على صف الرؤوس
    worksheet.auto_filter.ref = worksheet.dimensions

    apply_red_font_for_negatives(worksheet)
    # writer.close()
    # return output.getvalue()


